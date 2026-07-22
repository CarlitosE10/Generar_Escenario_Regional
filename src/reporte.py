"""Generación de reportes: Excel multi-hoja (con resaltado), gráficas y logs.

Los notebooks orquestadores llaman estas funciones para exportar los resultados
de la comparación/regionalización y visualizarlos inline.
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Nombres de columna legibles para los reportes en español
NOMBRES_REPORTE = {
    "Parametro": "Parámetro",
    "Parameter": "Parámetro",
    "TECHNOLOGY": "TECHNOLOGY",
    "FUEL": "FUEL",
    "Region": "Región",
    "Valor_Nacional": "Valor Nacional",
    "Valor_Regional": "Valor Regional",
    "Diferencia": "Diferencia",
    "Diferencia_Pct": "Diferencia %",
    "Tecnologia_Fuel": "Tecnología/Fuel",
    "Tipo_Anomalia": "Tipo_Anomalía",
}

_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FUENTE_ROJA = Font(color="9C0006")


def exportar_excel(path: str | Path, hojas: dict[str, pd.DataFrame], renombrar: bool = False) -> Path:
    """Escribe un Excel con una hoja por DataFrame (se omiten los vacíos).

    renombrar=True traduce los encabezados según NOMBRES_REPORTE.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    hojas_validas = {nombre: df for nombre, df in hojas.items() if df is not None and not df.empty}
    if not hojas_validas:
        raise ValueError("No hay hojas con datos para exportar")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for nombre, df in hojas_validas.items():
            salida = df.rename(columns=NOMBRES_REPORTE) if renombrar else df
            # Excel limita el nombre de hoja a 31 caracteres
            salida.to_excel(writer, sheet_name=nombre[:31], index=False)
    logger.info("Excel exportado: %s (%s hojas)", path, len(hojas_validas))
    return path


def exportar_reporte_comparacion(
    path: str | Path,
    comparacion: pd.DataFrame,
    anomalias: pd.DataFrame,
    resumen: pd.DataFrame,
    umbral_alerta_pct: float = 0.01,
    max_filas: int = 50_000,
    hojas_extra: dict[str, pd.DataFrame] | None = None,
) -> Path:
    """Reporte de comparación con hojas Comparacion / Anomalias / Resumen.

    En la hoja Comparacion, las filas con |Diferencia %| > umbral_alerta_pct
    se resaltan en rojo mediante UNA regla de formato condicional (no estilos
    por celda: con cientos de miles de filas resaltadas eso tardaba horas).
    Si la comparación supera max_filas se exportan solo las primeras (ya vienen
    ordenadas por |Diferencia| desc., así que se conservan las más relevantes).

    hojas_extra: {nombre_hoja: DataFrame} adicionales (p. ej. Inventario_Tech_Fuel,
    Por_Sector, Sector_Detalle, Anomalias_Contextualizadas); las vacías o None
    se omiten con warning.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if len(comparacion) > max_filas:
        logger.warning("Comparación con %s filas: se exportan las %s con mayor |Diferencia|",
                       f"{len(comparacion):,}", f"{max_filas:,}")
        comparacion = comparacion.head(max_filas)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        comparacion.rename(columns=NOMBRES_REPORTE).to_excel(writer, sheet_name="Comparacion", index=False)
        (anomalias if not anomalias.empty else pd.DataFrame([{"Info": "Sin anomalías"}])) \
            .rename(columns=NOMBRES_REPORTE).to_excel(writer, sheet_name="Anomalias", index=False)
        resumen.rename(columns=NOMBRES_REPORTE).to_excel(writer, sheet_name="Resumen", index=False)

        for nombre, df in (hojas_extra or {}).items():
            if df is None or df.empty:
                logger.warning("Hoja extra '%s' sin datos: se omite del reporte", nombre)
                continue
            df.rename(columns=NOMBRES_REPORTE).to_excel(writer, sheet_name=nombre[:31], index=False)

        # Resaltado en rojo: |Diferencia %| > umbral, como formato condicional
        # sobre todo el rango (la fila 1 es el encabezado). Excel evalúa la
        # regla al abrir el archivo; el resultado visual es el mismo que el
        # antiguo estilo por celda pero la escritura es O(1).
        if not comparacion.empty and "Diferencia_Pct" in comparacion.columns:
            ws = writer.book["Comparacion"]
            col_pct = get_column_letter(comparacion.columns.get_loc("Diferencia_Pct") + 1)
            rango = f"A2:{get_column_letter(len(comparacion.columns))}{len(comparacion) + 1}"
            ws.conditional_formatting.add(rango, FormulaRule(
                formula=[f"AND(ISNUMBER(${col_pct}2),ABS(${col_pct}2)>{umbral_alerta_pct})"],
                fill=_ROJO, font=_FUENTE_ROJA,
            ))
    logger.info("Reporte de comparación exportado: %s", path)
    return path


def resumen_discrepancias(discrepancias: pd.DataFrame, total_comparado: int, tolerancia: float) -> str:
    """Texto corto de estado para mostrar en el notebook orquestador."""
    if discrepancias.empty:
        return f"OK: todas las diferencias están dentro de la tolerancia ({tolerancia})."
    max_diff = discrepancias["Diferencia"].abs().max()
    return (
        f"ALERTA: {len(discrepancias):,} de {total_comparado:,} combinaciones superan "
        f"la tolerancia ({tolerancia}). Diferencia máxima: {max_diff:,.6f}"
    )


def top_diferencias(comparacion: pd.DataFrame, n: int = 20) -> pd.DataFrame:
    """Top-N de diferencias absolutas agregadas por tecnología o fuel.

    Agrupa por TECHNOLOGY si el parámetro la usa; si no, por FUEL. Suma
    |Diferencia| sobre los años para rankear.
    """
    if comparacion.empty:
        return pd.DataFrame()
    df = comparacion.copy()
    # Variable = TECHNOLOGY y, donde el parámetro no la usa, el FUEL
    tech = df["TECHNOLOGY"] if "TECHNOLOGY" in df.columns else pd.Series(pd.NA, index=df.index)
    fuel = df["FUEL"] if "FUEL" in df.columns else pd.Series(pd.NA, index=df.index)
    df["Variable"] = tech.fillna(fuel)
    if df["Variable"].isna().all():
        return pd.DataFrame()
    df["_abs"] = df["Diferencia"].abs()
    top = (
        df.groupby(["Parametro", "Variable"], dropna=False)
        .agg(Diferencia_Abs_Total=("_abs", "sum"), Max_Diferencia_Abs=("_abs", "max"),
             Filas=("_abs", "size"))
        .reset_index()
        .sort_values("Diferencia_Abs_Total", ascending=False)
        .head(n)
        .reset_index(drop=True)
    )
    return top.dropna(subset=["Variable"])


def grafica_top_diferencias(
    comparacion: pd.DataFrame,
    n: int = 20,
    titulo: str = "Top diferencias absolutas Nacional vs Regional",
    ruta_png: str | Path | None = None,
):
    """Barras agrupadas por parámetro: Top-N |Diferencia| por tecnología/fuel.

    Si se da ruta_png, exporta la gráfica como PNG junto al Excel.
    """
    top = top_diferencias(comparacion, n)
    if top.empty:
        print("Sin diferencias que graficar.")
        return None
    pivote = top.pivot_table(index="Variable", columns="Parametro",
                             values="Diferencia_Abs_Total", fill_value=0)
    pivote = pivote.loc[top["Variable"].drop_duplicates()[::-1]]  # mayor arriba
    fig, ax = plt.subplots(figsize=(11, max(3, 0.45 * len(pivote))))
    pivote.plot.barh(ax=ax)
    ax.set_xlabel("Suma de |Diferencia| (todos los años)")
    ax.set_title(titulo)
    ax.legend(title="Parámetro", fontsize=8)
    plt.tight_layout()
    if ruta_png is not None:
        ruta_png = Path(ruta_png)
        ruta_png.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(ruta_png, dpi=150, bbox_inches="tight")
        logger.info("Gráfica exportada: %s", ruta_png)
    return fig


def grafica_discrepancias_por_parametro(discrepancias: pd.DataFrame,
                                        titulo: str = "Discrepancias por parámetro"):
    """Barras horizontales: cuántas combinaciones fallan por parámetro."""
    if discrepancias.empty:
        print("Sin discrepancias que graficar.")
        return None
    col = "Parametro" if "Parametro" in discrepancias.columns else "Parameter"
    conteo = discrepancias.groupby(col).size().sort_values()
    fig, ax = plt.subplots(figsize=(10, max(2, 0.4 * len(conteo))))
    conteo.plot.barh(ax=ax)
    ax.set_xlabel("Combinaciones con diferencia")
    ax.set_title(titulo)
    plt.tight_layout()
    return fig


def grafica_participacion_regional(df_sand: pd.DataFrame, anio: str | int,
                                   col_dimension: str, titulo: str | None = None):
    """Barras apiladas del valor regionalizado por código base y región en un año.

    Espera una tabla SAND regionalizada (códigos '<prefijo>_<base>').
    """
    anio = str(anio)
    if df_sand.empty or anio not in df_sand.columns:
        print(f"Sin datos para el año {anio}.")
        return None
    df = df_sand[[col_dimension, anio]].copy()
    partes = df[col_dimension].astype(str).str.split("_", n=1, expand=True)
    df["Region"], df["Base"] = partes[0], partes[1]
    pivote = df.pivot_table(index="Base", columns="Region", values=anio,
                            aggfunc="sum", fill_value=0)
    fig, ax = plt.subplots(figsize=(11, max(3, 0.35 * len(pivote))))
    pivote.plot.barh(stacked=True, ax=ax)
    ax.set_xlabel("Valor regionalizado")
    ax.set_title(titulo or f"Distribución regional — {anio}")
    ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=8)
    plt.tight_layout()
    return fig


def escribir_log_regionalizacion(log: pd.DataFrame, carpeta: str | Path,
                                 nombre_base: str = "log_regionalizacion") -> tuple[Path, Path]:
    """Escribe el log de regionalización como Excel y texto plano, con timestamp."""
    carpeta = Path(carpeta)
    carpeta.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta_xlsx = carpeta / f"{nombre_base}_{stamp}.xlsx"
    ruta_txt = carpeta / f"{nombre_base}_{stamp}.txt"

    with pd.ExcelWriter(ruta_xlsx, engine="openpyxl") as writer:
        log.to_excel(writer, sheet_name="Log", index=False)

    lineas = [f"Log de regionalización — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "=" * 70]
    for _, fila in log.iterrows():
        contexto = " ".join(str(v) for v in (fila.get("TECHNOLOGY"), fila.get("FUEL")) if pd.notna(v))
        lineas.append(f"[{fila['Nivel']:<11}] {fila['Parametro']}"
                      + (f" ({contexto})" if contexto else "") + f": {fila['Mensaje']}")
    ruta_txt.write_text("\n".join(lineas), encoding="utf-8")
    logger.info("Log de regionalización: %s / %s", ruta_xlsx.name, ruta_txt.name)
    return ruta_xlsx, ruta_txt


def grafica_control_regionalizacion(
    df_sand: pd.DataFrame,
    parametro: str,
    indices: list[str],
    anios: list[str] | None = None,
    top_n: int = 10,
    max_paneles: int = 20,
):
    """Gráfica de control de un parámetro regionalizado: un panel por código base,
    barras apiladas por región a lo largo de los años.

    La dimensión de corte sale de los índices del parámetro (config_depurado.yaml):
    FUEL si el parámetro se indexa por FUEL (también cuando usa ambos, según la
    convención del proyecto), TECHNOLOGY en caso contrario. Si hay más de
    `max_paneles` códigos se muestran los `top_n` por valor total y el resto se
    agrega en un panel "Otros".

    Devuelve la figura, o None si no hay nada que graficar (el llamador avisa).
    """
    col = "FUEL" if "FUEL" in indices else "TECHNOLOGY"
    if df_sand is None or df_sand.empty or col not in df_sand.columns:
        return None

    cols_anio = [c for c in (anios or df_sand.columns) if str(c).isdigit() and c in df_sand.columns]
    if not cols_anio:
        return None

    df = df_sand[[col] + cols_anio].copy()
    partes = df[col].astype(str).str.split("_", n=1, expand=True)
    if partes.shape[1] < 2:
        return None
    df["Region"], df["Base"] = partes[0], partes[1]
    for c in cols_anio:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    if not df[cols_anio].notna().any().any():
        return None

    totales = df.groupby("Base")[cols_anio].sum().sum(axis=1).sort_values(ascending=False)
    if len(totales) > max_paneles:
        principales = list(totales.head(top_n).index)
        df["Base"] = df["Base"].where(df["Base"].isin(principales), "Otros")
        bases = principales + ["Otros"]
    else:
        bases = list(totales.index)

    n = len(bases)
    ncols = min(3, n)
    nfilas = (n + ncols - 1) // ncols
    fig, axes = plt.subplots(nfilas, ncols, figsize=(5.5 * ncols, 3.2 * nfilas),
                             squeeze=False, sharex=True)
    regiones = sorted(df["Region"].unique())
    colores = plt.cm.tab10.colors

    for pos, base in enumerate(bases):
        ax = axes[pos // ncols][pos % ncols]
        sub = df[df["Base"] == base]
        pivote = sub.groupby("Region")[cols_anio].sum().reindex(regiones).fillna(0)
        acumulado = None
        for i, region in enumerate(regiones):
            valores = pivote.loc[region].to_numpy(dtype=float)
            ax.bar(cols_anio, valores, bottom=acumulado, label=region,
                   color=colores[i % len(colores)], width=0.85)
            acumulado = valores if acumulado is None else acumulado + valores
        ax.set_title(base, fontsize=9)
        ax.tick_params(axis="x", labelrotation=90, labelsize=6)
        ax.tick_params(axis="y", labelsize=7)

    for pos in range(n, nfilas * ncols):      # apagar los ejes sobrantes
        axes[pos // ncols][pos % ncols].axis("off")

    manejadores, etiquetas = axes[0][0].get_legend_handles_labels()
    fig.legend(manejadores, etiquetas, loc="upper right", ncol=len(regiones), fontsize=8)
    fig.suptitle(f"{parametro} — regionalizado por {col} ({cols_anio[0]}–{cols_anio[-1]})",
                 fontsize=11, y=1.0)
    fig.tight_layout(rect=(0, 0, 1, 0.95))
    return fig
