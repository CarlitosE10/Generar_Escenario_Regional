"""Smoke tests de los módulos src/ contra los archivos reales del proyecto.

Ejecutar desde la raíz:  python tests/test_smoke.py
(no requiere pytest; termina con código != 0 si algo falla)
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pandas as pd

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ / "src"))

import comparador
import regionalizador
import reporte
import sand_io
import sincronizador
import utils
import yaml_parser

utils.configurar_logging()
FALLOS: list[str] = []


def check(condicion: bool, mensaje: str) -> None:
    estado = "OK " if condicion else "FALLO"
    print(f"[{estado}] {mensaje}")
    if not condicion:
        FALLOS.append(mensaje)


def main() -> None:
    tmp = Path(tempfile.mkdtemp(prefix="smoke_sand_"))

    # --- yaml_parser -----------------------------------------------------------
    cfg = yaml_parser.cargar_params_config(RAIZ / "config" / "params_config.yaml")
    # Rutas definidas localmente (ya no hay config/paths_config.yaml; ver la celda
    # "Configuración de rutas" de los notebooks 01/02/03).
    SAND_NACIONAL = RAIZ / "SAND_Nacional_base" / "scenario_CN_Nacional_166_Parameters_SAND_ORI.xlsx"
    SAND_REGIONAL = RAIZ / "SAND_Regional" / "SAND_Regional_Template.xlsx"
    CONFIG_OTOOLE = RAIZ / "template_config.yaml"
    otoole = yaml_parser.cargar_config_otoole(CONFIG_OTOOLE)
    # El config otoole es un template que sigue el patrón OSeMOSYS: puede traer
    # parámetros/sets de más (no usados) y crece con el tiempo (p. ej. el set
    # STORAGE). En vez de un conteo exacto y frágil, verificamos que cargue y
    # contenga lo que el flujo realmente consume.
    params_esenciales = {"AccumulatedAnnualDemand", "InputActivityRatio",
                         "TotalAnnualMaxCapacityInvestment",
                         "TotalTechnologyAnnualActivityLowerLimit"}
    sets_esenciales = {"TECHNOLOGY", "FUEL", "YEAR"}
    check(params_esenciales <= set(otoole["param"]) and sets_esenciales <= set(otoole["set"]),
          "yaml_parser: config otoole con los params/sets esenciales del flujo")
    check(len(cfg["prefijo_region"]) == 7, "yaml_parser: 7 regiones con prefijo")

    # --- utils ------------------------------------------------------------------
    check(utils.es_centinela(99999.0, cfg["valor_centinela"], True), "utils: 99999 es centinela")
    check(utils.es_centinela(9999, cfg["valor_centinela"], True), "utils: 9999 (patrón 9s) es centinela")
    check(not utils.es_centinela(12.5, cfg["valor_centinela"], True), "utils: 12.5 no es centinela")
    check(utils.split_regional_name("CA_TRABUS") == ("CA", "TRABUS"), "utils: split prefijo regional")
    check(utils.norm_indice(1.0) == "1", "utils: norm_indice 1.0 -> '1'")

    # --- sand_io ------------------------------------------------------------------
    df_nac = sand_io.cargar_sand(SAND_NACIONAL)
    df_reg = sand_io.cargar_sand(SAND_REGIONAL)
    anios = sand_io.columnas_anio(df_nac)
    # Los SAND concretos (y su tamaño) cambian según las rutas de arriba; no fijamos un
    # conteo exacto. Verificamos que cargaron con datos y la estructura esperada.
    check(len(df_nac) > 1000 and len(df_reg) > 1000
          and {"Parameter", "TECHNOLOGY", "FUEL"} <= set(df_nac.columns)
          and {"Parameter", "TECHNOLOGY", "FUEL"} <= set(df_reg.columns),
          "sand_io: SANDs cargados con datos y estructura SAND")
    check(anios[0] == "2022" and anios[-1] == "2055", "sand_io: años 2022–2055")

    # --- comparador: modo lista_parametros (aditivo + intensivo) ------------------
    res = comparador.comparar_escenarios(
        df_nac, df_reg, otoole["param"], cfg,
        modo="lista_parametros",
        parametros_filtro=["AccumulatedAnnualDemand", "CapitalCost"],
        fuels=["TRABUS", "TRAAVI"],
        tecnologias=["DEMINDNGSBOI"],
        modo_filtro="contiene",
    )
    comp = res["comparacion"]
    check(not comp.empty, "comparador: produce comparación")
    check((comp[comp["Parametro"] == "AccumulatedAnnualDemand"]["Region"] == "SUMA_REGIONES").all(),
          "comparador: aditivo compara contra SUMA_REGIONES")
    reg_int = set(comp[comp["Parametro"] == "CapitalCost"]["Region"])
    check(reg_int and reg_int.issubset(set(cfg["prefijo_region"].values()) | {"SIN_PREFIJO"}),
          "comparador: intensivo compara región por región")
    check(comp["Diferencia"].abs().is_monotonic_decreasing, "comparador: ordenado por |Diferencia| desc")
    check("Diferencia_Pct" in comp.columns, "comparador: columna Diferencia %")
    check(set(res["anomalias"].columns) == {"Parametro", "Tecnologia_Fuel", "Tipo_Anomalia", "Detalle"},
          "comparador: hoja de anomalías con columnas esperadas")
    check((res["resumen"].iloc[0]["Parametro"] == "== TOTAL =="), "comparador: resumen con fila total")

    # centinela: ninguna fila aditiva con valor nacional 9999/99999
    aad = comp[comp["Parametro"] == "AccumulatedAnnualDemand"]
    check(not aad["Valor_Nacional"].apply(lambda v: utils.es_centinela(v, cfg["valor_centinela"], True)).any(),
          "comparador: centinelas excluidos de la ruta aditiva")

    # anomalía global: código nacional ausente del regional en TODOS los
    # parámetros consultados (INDOTH_COA no está regionalizado)
    res_glob = comparador.comparar_escenarios(
        df_nac, df_reg, otoole["param"], cfg,
        modo="lista_parametros", parametros_filtro=["AccumulatedAnnualDemand"],
        fuels=["INDOTH_COA", "INDCLIM"], modo_filtro="exacto",
    )
    an_glob = res_glob["anomalias"]
    fila_glob = an_glob[an_glob["Tipo_Anomalia"] == "FUEL_NO_EN_REGIONAL"]
    check(len(fila_glob) == 1 and (fila_glob["Tecnologia_Fuel"] == "INDOTH_COA").all(),
          "comparador: anomalía global FUEL_NO_EN_REGIONAL (INDOTH_COA, no INDCLIM)")
    check((fila_glob["Parametro"] == "TODOS (consultados)").all(),
          "comparador: anomalía global marcada como 'TODOS (consultados)'")

    # las anomalías respetan los filtros de tecnología/fuel (bug corregido:
    # REGIONES_INCOMPLETAS se calculaba del lado regional sin filtrar)
    res_filtro = comparador.comparar_escenarios(
        df_nac, df_reg, otoole["param"], cfg,
        modo="lista_parametros",
        parametros_filtro=["AccumulatedAnnualDemand", "InputActivityRatio"],
        tecnologias=["RES"], fuels=["RES"], modo_filtro="contiene",
    )
    an_filtro = res_filtro["anomalias"]
    fuera = an_filtro[(an_filtro["Tecnologia_Fuel"] != "-")
                      & ~an_filtro["Tecnologia_Fuel"].astype(str).str.contains("RES")]
    check(fuera.empty, "comparador: las anomalías respetan el filtro TECHNOLOGY/FUEL")

    # parámetro sin YEAR (DiscountRate, indexado solo por REGION)
    res_ti = comparador.comparar_escenarios(
        df_nac, df_reg, otoole["param"], cfg,
        modo="parametro", parametros_filtro=["DiscountRate"],
    )
    check(len(res_ti["comparacion"]) >= 1 and res_ti["comparacion"]["Año"].isna().all(),
          "comparador: parámetro tiempo-independiente (DiscountRate) comparado sin Año")

    # --- comparador: inventario, análisis por sector y contextualización ----------
    regiones = list(cfg["prefijo_region"].values())
    mapeos = comparador.cargar_mapeos(RAIZ / "Insumos" / "Mapeo")
    inv = comparador.inventario_tech_fuel(
        df_nac, df_reg, regiones,
        diccionario_tech=mapeos["diccionario_tech"], diccionario_fuel=mapeos["diccionario_fuel"])
    # Los conteos deben cuadrar con los códigos únicos del nacional:
    # Solo Nacional + En ambos == universo TECHNOLOGY del SAND nacional
    tech = inv["tech"]
    n_solo_nac = (tech["Categoria"] == "Solo Nacional").sum()
    n_ambos = (tech["Categoria"] == "En ambos").sum()
    n_nac = df_nac["TECHNOLOGY"].dropna().apply(utils.normalize_text).nunique()
    check(n_solo_nac + n_ambos == n_nac,
          f"comparador: inventario TECHNOLOGY cuadra con el nacional ({n_solo_nac}+{n_ambos}=={n_nac})")
    check({"Estado_Mapeo", "Nombre_Equivalente", "Grupo"} <= set(tech.columns),
          "comparador: inventario enriquecido con el diccionario de mapeo")
    if mapeos["diccionario_fuel"] is not None:
        check(not inv["renombramientos_fuel"].empty
              and (inv["renombramientos_fuel"]["Fuente"] == "diccionario_fuel").all(),
              "comparador: renombramientos de FUEL desde diccionario_fuel.xlsx")

    check(comparador.detectar_sectores(df_nac) == comparador.SECTORES,
          "comparador: los 6 sectores activos en el nacional")
    df_reg_prep = comparador.preparar_regional(df_reg, regiones)
    sec = comparador.analisis_por_sector(
        df_nac, df_reg, "TRA", otoole["param"], cfg,
        diccionario_tech=mapeos["diccionario_tech"], diccionario_fuel=mapeos["diccionario_fuel"],
        df_reg_prep=df_reg_prep)
    rv = sec["resumen_validacion"]
    check(not rv.empty and (rv["Filas_OK"] + rv["Filas_Discrepancia"] == rv["Filas"]).all(),
          "comparador: análisis por sector TRA con conteos OK+discrepancia consistentes")
    check(sec["modos_tra"] is not None and not sec["modos_tra"].empty,
          "comparador: desglose por modo de transporte (MODOS_TRA)")
    desg = sec["desglose"]
    cols_reg = [c for c in regiones if c in desg.columns]
    # Solo filas con dato regional (las solo-nacionales dejan Suma_Regional en NaN)
    con_reg = desg[desg["Suma_Regional"].notna()]
    diff_desg = (con_reg[cols_reg].sum(axis=1) - con_reg["Suma_Regional"]).abs().max()
    check(not con_reg.empty and diff_desg < 1e-9,
          "comparador: desglose por región suma igual a Suma_Regional")
    sec_filtrado = comparador.analisis_por_sector(
        df_nac, df_reg, "TRA", otoole["param"], cfg,
        parametros=["AccumulatedAnnualDemand"], df_reg_prep=df_reg_prep)
    check(set(sec_filtrado["resumen_validacion"]["Parametro"]) == {"AccumulatedAnnualDemand"},
          "comparador: análisis por sector respeta el filtro de parámetros")
    check(not comparador.tabla_por_sector({"TRA": sec}).empty,
          "comparador: tabla pivot sector × parámetro")

    if mapeos["mapeo"] is not None:
        ctx = comparador.contextualizar_anomalias(res_glob["anomalias"], mapeos["mapeo"])
        check({"TIENE_MAPEO", "REGIONES_ESPERADAS", "NOMBRE_REGIONAL", "ACCION_SUGERIDA"}
              <= set(ctx.columns)
              and set(ctx["ACCION_SUGERIDA"]) <= {"cambio_de_nombre", "no_existe_en_region",
                                                  "verificar_creacion", "sin_info"},
              "comparador: anomalías contextualizadas con acción sugerida")

    # --- regionalizador ------------------------------------------------------------
    # Archivo industrial canónico fijo: la config de rutas activa puede
    # apuntar a otro escenario, pero la equivalencia de abajo es contra las
    # salidas industriales de referencia en SANDs_Reducidos/
    df_pct = regionalizador.cargar_participaciones(RAIZ / "config" / "participaciones_industriales.xlsx")
    check({"Parametro", "TECHNOLOGY", "FUEL", "Region", "Año", "Participacion"} <= set(df_pct.columns),
          "regionalizador: participaciones en formato largo estándar")
    malas = regionalizador.validar_participaciones(df_pct, cfg["tolerancia_participacion"])
    check(malas.empty, "regionalizador: participaciones suman 1.0")

    # Filtros explícitos derivados del archivo de participaciones: un filtro
    # amplio ('contiene IND') arrastraría códigos sin participación -> error
    fuels_aad = sorted(df_pct[df_pct["Parametro"] == "AccumulatedAnnualDemand"]["FUEL"].dropna().unique())
    techs_ll = sorted(df_pct[df_pct["Parametro"] == "TotalTechnologyAnnualActivityLowerLimit"]
                      ["TECHNOLOGY"].dropna().unique())
    res_aad = regionalizador.regionalizar(
        df_nac, df_reg, df_pct,
        parametros=["AccumulatedAnnualDemand"],
        params_otoole=otoole["param"], cfg=cfg,
        fuels_filtro=fuels_aad, modo_filtro="exacto",
    )
    res_ll = regionalizador.regionalizar(
        df_nac, df_reg, df_pct,
        parametros=["TotalTechnologyAnnualActivityLowerLimit"],
        params_otoole=otoole["param"], cfg=cfg,
        tecnologias_filtro=techs_ll, modo_filtro="exacto",
    )
    sands = {**res_aad["sands"], **res_ll["sands"]}
    log_total = pd.concat([res_aad["log"], res_ll["log"]], ignore_index=True)
    check(set(sands) == {"AccumulatedAnnualDemand", "TotalTechnologyAnnualActivityLowerLimit"},
          "regionalizador: genera SAND por parámetro")

    # --- flujo con mapeo: participación en una región donde el código NO existe ---
    # (motivo PARTICIPACION_REGION_INEXISTENTE; escenario sintético autocontenido)
    cols_v = utils.DIMENSION_COLS + ["2022", "2023"]
    nac_v = pd.DataFrame([{**{c: pd.NA for c in cols_v}, "Parameter": "AccumulatedAnnualDemand",
                           "REGION": "RE1", "FUEL": "TESTF", "2022": 100.0, "2023": 200.0}])
    reg_v = pd.DataFrame([{**{c: pd.NA for c in cols_v}, "Parameter": "AccumulatedAnnualDemand",
                           "REGION": "RE1", "FUEL": f"{p}_TESTF", "2022": 1.0} for p in ("AN", "CA")])
    pct_v = pd.DataFrame([{"Parametro": "AccumulatedAnnualDemand", "TECHNOLOGY": None, "FUEL": "TESTF",
                           "Region": r, "Año": a, "Participacion": v}
                          for a in (2022, 2023) for r, v in (("AN", 0.5), ("CA", 0.3), ("SE", 0.2))])
    mapeo_v = dict(rename_tech={}, rename_fuel={}, regiones_tech={}, regiones_fuel={},
                   obs_tech={}, obs_fuel={}, disponible=False)
    om_v = regionalizador.detectar_omisiones(nac_v, reg_v, pct_v, ["AccumulatedAnnualDemand"],
                                             otoole["param"], cfg, mapeo_v)
    check(list(om_v["Motivo"]) == [regionalizador.MOTIVO_PARTICIPACION_REGION_INEXISTENTE]
          and om_v.iloc[0]["Regiones_Faltantes"] == "SE"
          and om_v.iloc[0]["Decision"] == regionalizador.DECISION_CREAR_EXISTENTES,
          "regionalizador: detecta participación en región inexistente (SE) con default crear_existentes")
    sand_v = regionalizador.regionalizar_con_mapeo(
        nac_v, reg_v, pct_v, ["AccumulatedAnnualDemand"], otoole["param"], cfg, mapeo_v,
        omisiones=om_v)["sands"]["AccumulatedAnnualDemand"]
    check(set(sand_v["FUEL"]) == {"AN_TESTF", "CA_TESTF"},
          "regionalizador: por defecto NO emite la región inexistente (SE descartada)")
    om_crear = om_v.copy(); om_crear["Decision"] = "crear_regiones:SE"
    sand_c = regionalizador.regionalizar_con_mapeo(
        nac_v, reg_v, pct_v, ["AccumulatedAnnualDemand"], otoole["param"], cfg, mapeo_v,
        omisiones=om_crear)["sands"]["AccumulatedAnnualDemand"].set_index("FUEL")
    check("SE_TESTF" in sand_c.index and abs(float(sand_c.loc["SE_TESTF", "2022"]) - 20.0) < 1e-9,
          "regionalizador: decidir crear_regiones:SE emite SE con el % real del archivo (0.2*100=20)")

    # --- residencial IN/SE: reglas de enrutamiento (clasificar_destino) ---
    import residencial_in_se as res_inse
    ex_in = {"IN_RESCLIM", "IN_RESOTH", "IN_DEMRESELCAIR_PAR_HIG",
             "IN_DEMRESELCOTH_HIG", "IN_DEMRESELCOTH_MID", "IN_DEMRESELCOTH_LOW"}
    casos = [
        (("RESCLIM", "FUEL", "IN"), ("RESCLIM", res_inse.REGLA_COLAPSO)),
        (("RESILU", "FUEL", "IN"), ("RESOTH", res_inse.REGLA_FUEL_OTH)),
        (("DEMRESELCAIR_PAR_HIG", "TECHNOLOGY", "IN"), ("DEMRESELCAIR_PAR_HIG", res_inse.REGLA_COLAPSO)),
        (("DEMRESELCILU_HIG", "TECHNOLOGY", "IN"), ("DEMRESELCOTH_HIG", res_inse.REGLA_ELEC_OTH)),
        (("DEMRESELCTV_CRT", "TECHNOLOGY", "IN"), ("DEMRESELCOTH_MID", res_inse.REGLA_ELEC_OTH)),
    ]
    ok_routing = all(res_inse.clasificar_destino(b, d, p, ex_in) == esp for (b, d, p), esp in casos)
    check(ok_routing, "residencial_in_se: clasificar_destino aplica colapso / RESOTH / DEMRESELCOTH(efic)")
    check(res_inse.quitar_sufijo("DEMRESELCOTH_HIG_URB") == "DEMRESELCOTH_HIG"
          and res_inse.quitar_sufijo("IN_RESCLIM") == "IN_RESCLIM",
          "residencial_in_se: quitar_sufijo solo quita _URB/_RUR")

    # --- cargar_participaciones: columna 'Año' presente pero vacía = constante ---
    # Regresión: una 'Año' toda vacía hacía que se cargaran 0 filas (rama "por año"
    # filtraba Año.notna()), y TODO caía en SIN_PARTICIPACION.
    p_anio_vacio = pd.DataFrame({
        "Parameter": [None, None], "TECHNOLOGY": ["DEMTRADSLSHP", None], "FUEL": [None, "TRABUS"],
        "Año": [None, None], "CA": [0.6, 0.5], "OR": [0.1, 0.2], "SO": [0.1, 0.1],
        "AN": [0.1, 0.1], "NE": [0.05, 0.05], "SE": [0.03, 0.03], "IN": [0.02, 0.02]})
    ruta_pv = tmp / "part_anio_vacio.xlsx"
    p_anio_vacio.to_excel(ruta_pv, index=False)
    pct_pv = regionalizador.cargar_participaciones(
        ruta_pv, parametros=["ResidualCapacity", "AccumulatedAnnualDemand"], params_otoole=otoole["param"])
    check(len(pct_pv) > 0
          and not pct_pv[(pct_pv["Parametro"] == "ResidualCapacity")
                         & (pct_pv["TECHNOLOGY"] == "DEMTRADSLSHP")].empty,
          "regionalizador: cargar_participaciones trata columna 'Año' vacía como constante (no 0 filas)")

    # Regresión: un parámetro repetido en la lista no debe duplicar filas de
    # participación (si no, (Region, Año) repetido -> Series en pct_combo.get -> crash)
    pct_dup = regionalizador.cargar_participaciones(
        ruta_pv, parametros=["ResidualCapacity", "ResidualCapacity"], params_otoole=otoole["param"])
    check(not pct_dup.duplicated().any(),
          "regionalizador: cargar_participaciones deduplica filas con lista de parámetros repetidos")

    # Equivalencia con las salidas de los notebooks originales (SANDs_Reducidos)
    gen = sands["AccumulatedAnnualDemand"].set_index("FUEL")[anios].apply(
        pd.to_numeric, errors="coerce").sort_index()
    ref = sand_io.cargar_sand(RAIZ / "SANDs_Reducidos" / "SAND_AccumulatedAnnualDemand_Regional_Industrial.xlsx",
                              hoja=0, validar=False)
    ref = ref[ref["FUEL"].notna()]  # el archivo curado trae filas vacías
    cols = [c for c in anios if c in ref.columns]
    ref = ref.set_index("FUEL")[cols].astype(float).sort_index()
    comunes = gen.index.intersection(ref.index).unique()
    diff = (gen.loc[comunes, cols] - ref.loc[comunes]).abs().max().max()
    check(set(ref.index) <= set(gen.index) and diff < 1e-9,
          f"regionalizador: demanda reproduce el notebook original (Δmax={diff:.1e})")

    gen_l = sands["TotalTechnologyAnnualActivityLowerLimit"].set_index("TECHNOLOGY")[anios].apply(
        pd.to_numeric, errors="coerce").sort_index()
    ref_l = sand_io.cargar_sand(RAIZ / "SANDs_Reducidos" / "SAND_LowerLimit_Industria.xlsx", hoja=0, validar=False)
    ref_l = ref_l[ref_l["TECHNOLOGY"].notna()]
    cols_l = [c for c in anios if c in ref_l.columns]
    ref_l = ref_l.set_index("TECHNOLOGY")[cols_l].astype(float).sort_index()
    comunes_l = gen_l.index.intersection(ref_l.index).unique()
    diff_l = (gen_l.loc[comunes_l, cols_l] - ref_l.loc[comunes_l]).abs().max().max()
    check(set(ref_l.index) <= set(gen_l.index) and diff_l < 1e-9,
          f"regionalizador: límites reproducen el notebook original (Δmax={diff_l:.1e})")
    check((log_total["Nivel"] == "OK").sum() == 2, "regionalizador: log con 2 parámetros OK")

    # participación constante (columna 'Participacion') + centinela copiado
    pct_const = pd.DataFrame({
        "Parámetro": ["TotalAnnualMaxCapacityInvestment"] * 7,
        "TECHNOLOGY": ["DEMINDNGSBOI_LOW"] * 7,
        "FUEL": [None] * 7,
        "Región": list(cfg["prefijo_region"]),
        "Participacion": [0.3, 0.2, 0.2, 0.1, 0.1, 0.05, 0.05],
    })
    ruta_const = tmp / "pct_const.xlsx"
    pct_const.to_excel(ruta_const, index=False)
    df_pct_const = regionalizador.cargar_participaciones(ruta_const)
    check((df_pct_const["Año"] == regionalizador.ANIO_CONSTANTE).all(),
          "regionalizador: detecta participación constante")
    res_max = regionalizador.regionalizar(
        df_nac, df_reg, df_pct_const,
        parametros=["TotalAnnualMaxCapacityInvestment"],
        params_otoole=otoole["param"], cfg=cfg,
        tecnologias_filtro=["DEMINDNGSBOI_LOW"], modo_filtro="exacto",
    )
    sand_max = res_max["sands"]["TotalAnnualMaxCapacityInvestment"]
    fila_nac = df_nac[(df_nac["Parameter"] == "TotalAnnualMaxCapacityInvestment")
                      & (df_nac["TECHNOLOGY"] == "DEMINDNGSBOI_LOW")].iloc[0]
    anio_chk = next((a for a in anios if utils.es_centinela(fila_nac[a], cfg["valor_centinela"], True)), None)
    if anio_chk is not None:
        vals = pd.to_numeric(sand_max[anio_chk], errors="coerce").dropna().unique()
        check(len(vals) == 1 and utils.es_centinela(vals[0], cfg["valor_centinela"], True),
              "regionalizador: centinela se copia sin repartir")
    else:
        # sin centinelas en esta fila: verificar el reparto multiplicativo
        v = utils.safe_float(fila_nac[anios[0]]) or 0
        suma = pd.to_numeric(sand_max[anios[0]], errors="coerce").sum()
        check(abs(suma - v) < 1e-6, "regionalizador: reparto constante reproduce el nacional")

    # combo sin participación: se omite con log descriptivo y el flujo continúa
    res_sin_pct = regionalizador.regionalizar(
        df_nac, df_reg, df_pct,
        parametros=["ResidualCapacity"], params_otoole=otoole["param"], cfg=cfg,
        tecnologias_filtro=["DEMINDNGSBOI_LOW"], modo_filtro="exacto",
    )
    log_sin_pct = res_sin_pct["log"]
    omitidos = log_sin_pct[(log_sin_pct["Nivel"] == "OMITIDO")
                           & log_sin_pct["Mensaje"].str.contains("Participación no encontrada")]
    check(len(omitidos) == 1 and "ResidualCapacity" not in res_sin_pct["sands"],
          "regionalizador: combo sin participación se omite con log y el flujo continúa")

    # --- regionalizador: crear_existentes y atajo "aplicar a todos" por motivo -------
    # resolver_omisiones: [5]+lote solo para REGIONES_INCOMPLETAS; otro motivo no hereda
    def _om_fila(param, fuel, motivo, existentes="AN,CA", faltantes="OR"):
        return {"Parametro": param, "TECHNOLOGY": None, "FUEL": fuel, "Motivo": motivo,
                "Regiones_Existentes": existentes, "Regiones_Faltantes": faltantes,
                "Nombre_Regional": "", "Observacion": "",
                "Decision": regionalizador.DECISION_OMITIR}

    om_test = pd.DataFrame([
        _om_fila("AccumulatedAnnualDemand", "F1", regionalizador.MOTIVO_REGIONES_INCOMPLETAS),
        _om_fila("AccumulatedAnnualDemand", "F2", regionalizador.MOTIVO_REGIONES_INCOMPLETAS),
        _om_fila("AccumulatedAnnualDemand", "F3", regionalizador.MOTIVO_SIN_PARTICIPACION),
    ])
    # F1: [5] + aplicar a todos; F2 hereda; F3: omitir sin lote
    respuestas = iter(["5", "s", "3", "n"])
    om_res = regionalizador.resolver_omisiones(
        om_test, interactivo=True, entrada=lambda _: next(respuestas), salida=lambda *_: None)
    check(om_res["Decision"].tolist() == [regionalizador.DECISION_CREAR_EXISTENTES,
                                          regionalizador.DECISION_CREAR_EXISTENTES,
                                          regionalizador.DECISION_OMITIR],
          "regionalizador: [5]+lote aplica crear_existentes a todo el motivo sin tocar otros")

    # [5] en SIN_PARTICIPACION = crear en las existentes con reparto 1/N (+lote)
    om_test2 = pd.DataFrame([
        _om_fila("AccumulatedAnnualDemand", "F1", regionalizador.MOTIVO_SIN_PARTICIPACION),
        _om_fila("AccumulatedAnnualDemand", "F2", regionalizador.MOTIVO_SIN_PARTICIPACION),
        _om_fila("AccumulatedAnnualDemand", "F3", regionalizador.MOTIVO_SIN_PARTICIPACION),
    ])
    respuestas2 = iter(["5", "s"])       # F1: [5] + aplicar a todos; F2 y F3 heredan
    om_res2 = regionalizador.resolver_omisiones(
        om_test2, interactivo=True, entrada=lambda _: next(respuestas2), salida=lambda *_: None)
    check(om_res2["Decision"].tolist() == [regionalizador.DECISION_CREAR_EXISTENTES_UNIF] * 3,
          "regionalizador: [5] en SIN_PARTICIPACION crea en las existentes (1/N) y se aplica en lote")

    # La [3] ("no crear") también ofrece el lote: en PARTICIPACION_REGION_INEXISTENTE
    # su decisión es crear_existentes, no omitir. Otro motivo no lo hereda.
    om_test3 = pd.DataFrame([
        _om_fila("AccumulatedAnnualDemand", "F1",
                 regionalizador.MOTIVO_PARTICIPACION_REGION_INEXISTENTE),
        _om_fila("AccumulatedAnnualDemand", "F2",
                 regionalizador.MOTIVO_PARTICIPACION_REGION_INEXISTENTE),
        _om_fila("AccumulatedAnnualDemand", "F3", regionalizador.MOTIVO_SIN_CORRESPONDENCIA),
    ])
    respuestas3 = iter(["3", "s", "1", "n"])   # F1: [3]+lote; F2 hereda; F3: [1] sin lote
    om_res3 = regionalizador.resolver_omisiones(
        om_test3, interactivo=True, entrada=lambda _: next(respuestas3), salida=lambda *_: None)
    check(om_res3["Decision"].tolist() == [regionalizador.DECISION_CREAR_EXISTENTES,
                                           regionalizador.DECISION_CREAR_EXISTENTES,
                                           regionalizador.DECISION_CREAR_TODAS],
          "regionalizador: [3]+lote aplica el 'no crear' del motivo sin tocar otros motivos")

    # regionalizar_con_mapeo con crear_existentes: % tal cual, sin renormalizar.
    # Nacional=100, mapeo espera AN,CA,OR pero solo existen AN,CA;
    # pct AN=0.5/CA=0.3/OR=0.2 -> AN=50, CA=30, sin fila OR (suma 80, a propósito)
    cols_syn = ["Parameter", "REGION", "TECHNOLOGY", "EMISSION", "MODE_OF_OPERATION",
                "FUEL", "TIMESLICE", "STORAGE", "REGION2", utils.TIME_INDEP_COL,
                "2022", "2023"]
    df_nac_syn = pd.DataFrame([{"Parameter": "AccumulatedAnnualDemand", "REGION": "RE1",
                                "FUEL": "F1", "2022": 100.0, "2023": 100.0}],
                              columns=cols_syn)
    df_reg_syn = pd.DataFrame([{"Parameter": "AccumulatedAnnualDemand", "REGION": "RE1",
                                "FUEL": f"{p}_F1"} for p in ("AN", "CA")], columns=cols_syn)
    mapeo_syn = {"regiones_tech": {}, "regiones_fuel": {"F1": ["AN", "CA", "OR"]},
                 "rename_tech": {}, "rename_fuel": {}, "obs_tech": {}, "obs_fuel": {},
                 "disponible": True}
    pct_syn = pd.DataFrame([
        {"Parametro": "AccumulatedAnnualDemand", "TECHNOLOGY": None, "FUEL": "F1",
         "Region": r, "Año": a, "Participacion": p}
        for a in (2022, 2023) for r, p in (("AN", 0.5), ("CA", 0.3), ("OR", 0.2))])
    om_syn = pd.DataFrame([{**_om_fila("AccumulatedAnnualDemand", "F1",
                                       regionalizador.MOTIVO_REGIONES_INCOMPLETAS),
                            "Decision": regionalizador.DECISION_CREAR_EXISTENTES}])
    res_ce = regionalizador.regionalizar_con_mapeo(
        df_nac_syn, df_reg_syn, pct_syn, parametros=["AccumulatedAnnualDemand"],
        params_otoole=otoole["param"], cfg=cfg, mapeo=mapeo_syn, omisiones=om_syn)
    sand_ce = res_ce["sands"].get("AccumulatedAnnualDemand", pd.DataFrame(columns=cols_syn))
    vals_ce = pd.to_numeric(sand_ce.set_index("FUEL")["2022"], errors="coerce")
    check(set(sand_ce["FUEL"]) == {"AN_F1", "CA_F1"}
          and abs(vals_ce.get("AN_F1", 0) - 50.0) < 1e-9
          and abs(vals_ce.get("CA_F1", 0) - 30.0) < 1e-9,
          "regionalizador: crear_existentes reparte tal cual (AN=50, CA=30, sin OR)")
    check(res_ce["log"]["Mensaje"].str.contains("Emitido solo en regiones existentes").any()
          and int(res_ce["resumen"].iloc[0]["Combos_Creados"]) == 1,
          "regionalizador: crear_existentes queda registrado como CREADO en log y resumen")

    # Sin % aplicable a las existentes: no se inventa 1/N -> ADVERTENCIA y sin filas
    pct_syn_or = pct_syn[pct_syn["Region"] == "OR"].copy()   # solo la región faltante
    res_ce2 = regionalizador.regionalizar_con_mapeo(
        df_nac_syn, df_reg_syn, pct_syn_or, parametros=["AccumulatedAnnualDemand"],
        params_otoole=otoole["param"], cfg=cfg, mapeo=mapeo_syn, omisiones=om_syn)
    check("AccumulatedAnnualDemand" not in res_ce2["sands"]
          and res_ce2["log"]["Mensaje"].str.contains(
              "crear_existentes sin participación aplicable").any(),
          "regionalizador: crear_existentes sin % aplicable no emite (ADVERTENCIA, sin 1/N)")

    # crear_existentes_uniforme (SIN_PARTICIPACION): sin % en el archivo, 1/N entre
    # las existentes -> AN=50, CA=50 (nada en OR, donde el código no existe)
    pct_otro = pct_syn.assign(FUEL="F9")          # participaciones de otro combo: F1 sin %
    om_unif = pd.DataFrame([{**_om_fila("AccumulatedAnnualDemand", "F1",
                                        regionalizador.MOTIVO_SIN_PARTICIPACION),
                             "Decision": regionalizador.DECISION_CREAR_EXISTENTES_UNIF}])
    res_cu = regionalizador.regionalizar_con_mapeo(
        df_nac_syn, df_reg_syn, pct_otro, parametros=["AccumulatedAnnualDemand"],
        params_otoole=otoole["param"], cfg=cfg, mapeo=mapeo_syn, omisiones=om_unif)
    sand_cu = res_cu["sands"].get("AccumulatedAnnualDemand", pd.DataFrame(columns=cols_syn))
    vals_cu = pd.to_numeric(sand_cu.set_index("FUEL")["2022"], errors="coerce")
    check(set(sand_cu["FUEL"]) == {"AN_F1", "CA_F1"}
          and abs(vals_cu.get("AN_F1", 0) - 50.0) < 1e-9
          and abs(vals_cu.get("CA_F1", 0) - 50.0) < 1e-9
          and res_cu["log"]["Mensaje"].str.contains("uniforme 1/2").any(),
          "regionalizador: crear_existentes_uniforme reparte 1/N entre las existentes")

    # Participación toda-0 (fila-plantilla de 00_Generar_Mapeo) = sin dato: en un
    # condicional los años 0 del nacional se fijan y los reales quedan VACÍOS
    # (regresión: antes se multiplicaba por 0 y un "sin dato" se volvía límite 0)
    df_nac_cond = pd.DataFrame([{"Parameter": "TotalAnnualMaxCapacityInvestment",
                                 "REGION": "RE1", "TECHNOLOGY": "T1",
                                 "2022": 0.0, "2023": 5.0}], columns=cols_syn)
    df_reg_cond = pd.DataFrame([{"Parameter": "TotalAnnualMaxCapacityInvestment",
                                 "REGION": "RE1", "TECHNOLOGY": f"{p}_T1"}
                                for p in ("AN", "CA")], columns=cols_syn)
    mapeo_cond = {"regiones_tech": {"T1": ["AN", "CA"]}, "regiones_fuel": {},
                  "rename_tech": {}, "rename_fuel": {}, "obs_tech": {}, "obs_fuel": {},
                  "disponible": True}
    pct_cero = pd.DataFrame([
        {"Parametro": "TotalAnnualMaxCapacityInvestment", "TECHNOLOGY": "T1",
         "FUEL": None, "Region": r, "Año": a, "Participacion": 0.0}
        for a in (2022, 2023) for r in ("AN", "CA")])
    res_cond = regionalizador.regionalizar_con_mapeo(
        df_nac_cond, df_reg_cond, pct_cero,
        parametros=["TotalAnnualMaxCapacityInvestment"],
        params_otoole=otoole["param"], cfg=cfg, mapeo=mapeo_cond)
    sand_cond = res_cond["sands"]["TotalAnnualMaxCapacityInvestment"]
    v22 = pd.to_numeric(sand_cond["2022"], errors="coerce")
    check(len(sand_cond) == 2 and (v22 == 0).all() and sand_cond["2023"].isna().all()
          and res_cond["log"]["Mensaje"].str.contains(
              "Condicional con límites reales sin participación").any(),
          "regionalizador: participación toda-0 = sin dato (años 0 fijados, reales vacíos)")

    # --- regionalizador: participaciones.xlsx, único insumo de participaciones --------
    # Formato ancho Parameter|TECHNOLOGY|FUEL|Año|<prefijos>: el parámetro se infiere de
    # la lista y de cada uno se descartan las filas de la dimensión que no lo indexa
    # (SpecifiedAnnualDemand no se indexa por TECHNOLOGY, LowerLimit no por FUEL).
    ruta_pct = RAIZ / "Insumos" / "Mapeo" / "participaciones.xlsx"
    if ruta_pct.exists():
        pct_canon = regionalizador.cargar_participaciones(
            ruta_pct,
            parametros=["SpecifiedAnnualDemand", "TotalTechnologyAnnualActivityLowerLimit"],
            params_otoole=otoole["param"])
        pct_sad = pct_canon[pct_canon["Parametro"] == "SpecifiedAnnualDemand"]
        pct_low = pct_canon[pct_canon["Parametro"] == "TotalTechnologyAnnualActivityLowerLimit"]
        check(not pct_sad.empty and pct_sad["FUEL"].notna().all()
              and pct_sad["TECHNOLOGY"].isna().all()
              and not pct_low.empty and pct_low["TECHNOLOGY"].notna().all()
              and set(pct_canon["Region"].unique()) <= set(utils.REGIONES),
              "regionalizador: participaciones.xlsx ancho -> largo (índice no aplicable descartado)")

        # El reparto aditivo aplica el % del archivo región por región
        fuel_res = sorted(pct_sad["FUEL"].unique())[0]
        res_fr = regionalizador.regionalizar(
            df_nac, df_reg, pct_sad, parametros=["SpecifiedAnnualDemand"],
            params_otoole=otoole["param"], cfg=cfg,
            fuels_filtro=[fuel_res], modo_filtro="exacto")
        sand_fr = res_fr["sands"]["SpecifiedAnnualDemand"].set_index("FUEL")
        nac_fr = float(df_nac[(df_nac["Parameter"] == "SpecifiedAnnualDemand")
                              & (df_nac["FUEL"] == fuel_res)].iloc[0]["2030"])
        pct_an = float(pct_sad[(pct_sad["FUEL"] == fuel_res) & (pct_sad["Region"] == "AN")
                               & (pct_sad["Año"] == 2030)].iloc[0]["Participacion"])
        val_an = float(pd.to_numeric(sand_fr.loc[f"AN_{fuel_res}", "2030"], errors="coerce"))
        check(abs(val_an - nac_fr * pct_an) < 1e-9,
              f"regionalizador: reparto desde participaciones.xlsx = nacional × % ({fuel_res}, AN)")

        # residencial_in_se consume el MISMO archivo (no los Insumos/Participacion_*)
        anios_res = [int(c) for c in anios if int(c) <= 2054]
        dim_sad = res_inse.dimension_parametro("SpecifiedAnnualDemand", otoole["param"])
        pct_inse = res_inse.cargar_participacion(
            ruta_pct, "SpecifiedAnnualDemand", dim_sad,
            params_otoole=otoole["param"], anios=anios_res)
        check(dim_sad == "FUEL" and not pct_inse.empty
              and set(pct_inse["prefijo"].unique()) <= {"IN", "SE"}
              and set(pct_inse["Anio"]) <= set(anios_res),
              "residencial_in_se: cargar_participacion lee participaciones.xlsx y acota a IN/SE")
    else:
        print("       (omitido: Insumos/Mapeo/participaciones.xlsx no existe)")

    # Regiones anchas constantes con prefijos AN..SO y alias 'Fuel'/'Tecnologia'
    ruta_ej1 = RAIZ / "Insumos" / "ejemplo1.xlsx"
    if ruta_ej1.exists():
        pct_ej1 = regionalizador.cargar_participaciones(ruta_ej1, params_otoole=otoole["param"])
        check((pct_ej1["Año"] == regionalizador.ANIO_CONSTANTE).all()
              and set(pct_ej1["Region"].unique()) <= set(utils.REGIONES),
              "regionalizador: formato ancho constante con prefijos AN..SO")
    else:
        print("       (omitido: Insumos/ejemplo1.xlsx no existe)")

    # Solo regiones (comodín '*'): aplica a cualquier combo del parámetro
    ruta_ej3 = RAIZ / "Insumos" / "ejemplo3.xlsx"
    if ruta_ej3.exists():
        pct_ej3 = regionalizador.cargar_participaciones(
            ruta_ej3, parametros=["AccumulatedAnnualDemand"], params_otoole=otoole["param"])
        check((pct_ej3["TECHNOLOGY"] == regionalizador.COMODIN).all(),
              "regionalizador: archivo solo-regiones produce participación comodín '*'")
        res_ej3 = regionalizador.regionalizar(
            df_nac, df_reg, pct_ej3, parametros=["AccumulatedAnnualDemand"],
            params_otoole=otoole["param"], cfg=cfg,
            fuels_filtro=["INDDHT"], modo_filtro="exacto")
        sand_ej3 = res_ej3["sands"]["AccumulatedAnnualDemand"]
        nac_dht = float(df_nac[(df_nac["Parameter"] == "AccumulatedAnnualDemand")
                               & (df_nac["FUEL"] == "INDDHT")].iloc[0]["2030"])
        suma_ej3 = pd.to_numeric(sand_ej3["2030"], errors="coerce").sum()
        check(abs(suma_ej3 - nac_dht) < 1e-6,
              "regionalizador: comodín reparte manteniendo el código del nacional")
    else:
        print("       (omitido: Insumos/ejemplo3.xlsx no existe)")

    # Sin columna Parámetro y sin lista -> error claro
    if ruta_fuel_res.exists():
        try:
            regionalizador.cargar_participaciones(ruta_fuel_res)
            check(False, "regionalizador: archivo sin 'Parámetro' debe exigir la lista")
        except ValueError as exc:
            check("PARAMETROS_A_REGIONALIZAR" in str(exc),
                  "regionalizador: error claro si falta 'Parámetro' y no se pasa la lista")

    # --- sincronizador (flujo integrado, notebooks/03) ---------------------------------
    mapeo_reg = regionalizador.cargar_mapeo_regional(RAIZ / "Insumos" / "Mapeo")
    params_sync = ["AccumulatedAnnualDemand", "CapitalCost"]
    res_sync = comparador.comparar_escenarios(
        df_nac, df_reg, otoole["param"], cfg,
        modo="lista_parametros", parametros_filtro=params_sync)

    resumen_dif = sincronizador.resumen_diferencias(res_sync["comparacion"], umbral_pct=1.0)
    check(set(resumen_dif["Parametro"]) == set(params_sync)
          and set(resumen_dif["Estado"]) <= {"OK", "REVISAR"},
          "sincronizador: resumen de diferencias por parámetro")

    ruta_pct_sync = RAIZ / "Insumos" / "Mapeo" / "participaciones.xlsx"
    pct_sync = (regionalizador.cargar_participaciones(
        ruta_pct_sync, parametros=params_sync, params_otoole=otoole["param"])
        if ruta_pct_sync.exists() else None)

    dec = sincronizador.construir_decisiones(
        res_sync["comparacion"], res_sync["anomalias"], df_nac, df_reg,
        mapeo_reg, cfg, umbral_pct=1.0, df_pct=pct_sync)
    check(not dec.empty and list(dec.columns) == sincronizador.COLUMNAS_DECISIONES,
          "sincronizador: tabla de decisiones con las columnas esperadas")
    check(set(dec["ACCION"]) <= sincronizador.ACCIONES_VALIDAS,
          "sincronizador: toda ACCION inferida es válida")
    # Un renombramiento conocido (ELC -> ELC003) no debe proponerse para regionalizar:
    # la diferencia es de nomenclatura. Regresión del bug de None->NaN en apply().
    renombradas = dec[dec["Nombre_Regional"].astype(str).str.strip().ne("")
                      & dec["Nombre_Regional"].notna()]
    check(not renombradas.empty
          and not renombradas["Nombre_Regional"].astype(str).eq("nan").any()
          and (renombradas["ACCION"] == sincronizador.ACCION_IGNORAR).all(),
          "sincronizador: renombramientos detectados con su nombre real y marcados 'ignorar'")

    # Un código que existe en el regional no puede clasificarse como "faltante":
    # SIN_CORRESPONDENCIA es por parámetro, no por código.
    faltantes = dec[dec["TIPO_DIFERENCIA"].isin(
        [sincronizador.TIPO_TECH_FALTANTE, sincronizador.TIPO_FUEL_FALTANTE])]
    check(faltantes["Regiones_Existentes"].astype(str).str.strip().eq("").all(),
          "sincronizador: 'faltante' solo si el código no existe en ninguna región")
    sin_datos = dec[dec["TIPO_DIFERENCIA"] == sincronizador.TIPO_SIN_DATOS_PARAMETRO]
    check(not sin_datos.empty
          and sin_datos["Regiones_Existentes"].astype(str).str.strip().ne("").all(),
          "sincronizador: código existente sin datos del parámetro se reclasifica")

    # Cobertura parcial que coincide con lo que el mapeo espera = falso positivo
    # (solo aplica si el código existe en alguna región; si no existe en ninguna
    # es una ausencia real y se resuelve por la rama de faltantes)
    incompletas = dec[(dec["TIPO_DIFERENCIA"] == sincronizador.TIPO_REGIONES_INCOMPLETAS)
                      & dec["Regiones_Faltantes"].astype(str).str.strip().eq("")
                      & dec["Regiones_Existentes"].astype(str).str.strip().ne("")
                      & dec["Nombre_Regional"].astype(str).str.strip().eq("")]
    check(not incompletas.empty
          and (incompletas["ACCION"] == sincronizador.ACCION_IGNORAR).all(),
          "sincronizador: cobertura parcial esperada por el mapeo se marca 'ignorar'")

    sin_regla = dec["Motivo_Accion"].str.startswith("Sin regla aplicable").sum()
    check(sin_regla == 0,
          f"sincronizador: ninguna diferencia cae al fallback sin regla ({sin_regla})")

    ruta_dec = tmp / "decisiones_pendientes.xlsx"
    sincronizador.exportar_decisiones(dec, ruta_dec)
    dec_leida = sincronizador.leer_decisiones(ruta_dec)
    check(len(dec_leida) == len(dec)
          and dec_leida["ACCION"].tolist() == dec["ACCION"].tolist(),
          "sincronizador: ida y vuelta por Excel conserva las decisiones")

    dec_mala = dec.head(3).copy()
    dec_mala["ACCION"] = "accion_inventada"
    sincronizador.exportar_decisiones(dec_mala, tmp / "dec_mala.xlsx")
    try:
        sincronizador.leer_decisiones(tmp / "dec_mala.xlsx")
        check(False, "sincronizador: ACCION inválida debe abortar en modo estricto")
    except ValueError as exc:
        check("no reconocida" in str(exc), "sincronizador: ACCION inválida aborta con error claro")
    tolerante = sincronizador.leer_decisiones(tmp / "dec_mala.xlsx", estricto=False)
    check((tolerante["ACCION"] == sincronizador.ACCION_MANTENER).all(),
          "sincronizador: modo no estricto degrada la ACCION inválida a mantener_regional")

    if pct_sync is not None:
        aplicado = sincronizador.aplicar_decisiones(
            df_nac, df_reg, pct_sync, dec_leida, otoole["param"], cfg, mapeo_reg,
            years_filtro=list(range(2022, 2055)))
        n_reg = (dec_leida["ACCION"] == sincronizador.ACCION_REGIONALIZAR).sum()
        check(bool(aplicado["sands"]) == (n_reg > 0),
              f"sincronizador: aplica las {n_reg} decisiones 'regionalizar'")
        # Solo se tocan los combos decididos: nada fuera de la lista
        decididos = {(utils.normalize_text(f["TECHNOLOGY"]), utils.normalize_text(f["FUEL"]))
                     for _, f in dec_leida[dec_leida["ACCION"]
                                           == sincronizador.ACCION_REGIONALIZAR].iterrows()}
        bases = set()
        for df_s in aplicado["sands"].values():
            for col in ("TECHNOLOGY", "FUEL"):
                bases |= {utils.split_regional_name(v, regiones)[1]
                          for v in df_s[col].dropna().unique()}
        esperados = {c for combo in decididos for c in combo if c}
        check(bases <= esperados,
              "sincronizador: la regionalización se limita a los combos decididos")

        reg_sync = sincronizador.integrar_sands(df_reg, aplicado["sands"])
        filas_nuevas = sum(len(d) for d in aplicado["sands"].values())
        check(len(reg_sync) >= len(df_reg) and len(reg_sync) <= len(df_reg) + filas_nuevas,
              "sincronizador: el SAND sincronizado sustituye filas en vez de duplicarlas")

        res_desp = comparador.comparar_escenarios(
            df_nac, reg_sync, otoole["param"], cfg,
            modo="lista_parametros", parametros_filtro=params_sync)
        resolucion = sincronizador.comparar_resolucion(
            res_sync["comparacion"], res_desp["comparacion"], dec_leida, umbral_pct=1.0)
        m = resolucion["metricas"]
        check(m["resueltas"] > 0 and m["nuevas"] == 0,
              f"sincronizador: {m['resueltas']} diferencias resueltas, {m['nuevas']} nuevas")
        check(set(resolucion["detalle"]["Estado"]) <= {"RESUELTA", "PERSISTE", "NUEVA"},
              "sincronizador: estados de resolución válidos")
    else:
        print("       (omitido: Insumos/Mapeo/participaciones.xlsx no existe)")

    # --- reporte ---------------------------------------------------------------------
    ruta_rep = reporte.exportar_reporte_comparacion(
        tmp / "reporte_comp.xlsx", comp, res["anomalias"], res["resumen"],
        umbral_alerta_pct=cfg["umbral_alerta_pct"],
    )
    check(ruta_rep.exists(), "reporte: Excel de comparación con 3 hojas")
    ruta_ext = reporte.exportar_reporte_comparacion(
        tmp / "reporte_ext.xlsx", comp, res["anomalias"], res["resumen"],
        umbral_alerta_pct=cfg["umbral_alerta_pct"],
        hojas_extra={"Sector_Detalle": rv, "Vacia": pd.DataFrame()},
    )
    from openpyxl import load_workbook
    hojas_ext = load_workbook(ruta_ext, read_only=True).sheetnames
    check("Sector_Detalle" in hojas_ext and "Vacia" not in hojas_ext,
          "reporte: hojas_extra se agregan y las vacías se omiten")
    top = reporte.top_diferencias(comp, n=5)
    check(len(top) <= 5 and "Diferencia_Abs_Total" in top.columns, "reporte: top-N de diferencias")
    fig = reporte.grafica_top_diferencias(comp, n=5, ruta_png=tmp / "top.png")
    check(fig is not None and (tmp / "top.png").exists(), "reporte: gráfica top-N exportada a PNG")
    rutas_sand = regionalizador.escribir_sands(sands, tmp, "SmokeTest")
    check(len(rutas_sand) == 2 and all(r.exists() for r in rutas_sand), "reporte: SANDs reducidos escritos")
    ruta_log = reporte.escribir_log_regionalizacion(log_total, tmp)
    check(all(r.exists() for r in ruta_log), "reporte: log de regionalización (xlsx + txt)")

    print(f"\n{'='*60}")
    if FALLOS:
        print(f"{len(FALLOS)} FALLOS:")
        for f in FALLOS:
            print(f"  - {f}")
        sys.exit(1)
    print("TODOS LOS SMOKE TESTS OK")


if __name__ == "__main__":
    main()
